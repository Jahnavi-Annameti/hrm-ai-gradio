"""
HRM AI Business Automation Agent — Gradio Edition
--------------------------------------------------
Single-process Gradio app that replicates every feature of the original
FastAPI + React project (resume screening, AI decisioning, candidate
pipeline management, Excel exports, JD generator, interview question
generator, and email sending) without needing a separate backend server.

Run:
    pip install -r requirements.txt
    python app.py
"""

import os
import re
import io
import sqlite3
import smtplib
from datetime import datetime
from email.mime.text import MIMEText

import gradio as gr
import pandas as pd

# --------------------------------------------------------------------------
# Paths / DB setup
# --------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)

DB_PATH = os.path.join(DATA_DIR, "hrm_ai.db")
SELECTED_EXCEL_PATH = os.path.join(DATA_DIR, "selected_candidates.xlsx")
WAITING_EXCEL_PATH = os.path.join(DATA_DIR, "waiting_list_candidates.xlsx")
REJECTED_EXCEL_PATH = os.path.join(DATA_DIR, "rejected_candidates.xlsx")
ALL_EXCEL_PATH = os.path.join(DATA_DIR, "all_candidates.xlsx")


def db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = db()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS candidates(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            email TEXT,
            role TEXT,
            score INTEGER,
            decision TEXT,
            status TEXT,
            date TEXT,
            email_body TEXT,
            experience TEXT,
            qualification TEXT,
            applicant_skills TEXT,
            missing_skills TEXT,
            reason TEXT,
            fit_level TEXT,
            projects TEXT,
            project_links TEXT
        )
        """
    )
    conn.commit()
    conn.close()


init_db()

# --------------------------------------------------------------------------
# Static option lists (parity with the original React frontend)
# --------------------------------------------------------------------------
JOB_ROLES = [
    "Marketing Intern", "HR Intern", "Software Developer", "Frontend Developer",
    "Backend Developer", "Data Analyst", "Data Scientist", "Machine Learning Intern",
    "AI Intern", "UI UX Designer", "Graphic Designer", "Digital Marketing Executive",
    "SEO Executive", "Content Writer", "Social Media Manager", "Business Analyst",
    "Sales Executive", "Customer Support Executive", "Project Coordinator",
    "Finance Intern", "Operations Executive", "Recruiter", "Talent Acquisition Associate",
    "Product Manager Intern",
]

QUALIFICATION_OPTIONS = [
    "Any Degree", "B.Tech", "B.E", "B.Sc", "B.Com", "BBA", "MBA", "MCA", "M.Tech",
    "Diploma", "Relevant Certification", "Marketing / Communication Background",
]

EXPERIENCE_OPTIONS = [
    "<0 Years (Fresher)", "<1 Year", "1-2 Years", "2-3 Years", "3-5 Years", "5+ Years",
]

SKILL_BANK = [
    "python", "java", "c++", "javascript", "react", "node", "fastapi", "sql",
    "excel", "power bi", "tableau", "machine learning", "deep learning",
    "data analysis", "data visualization", "communication", "canva", "seo",
    "content writing", "social media marketing", "marketing", "presentation",
    "teamwork", "leadership", "project management", "html", "css", "mongodb",
    "sqlite", "pandas", "numpy", "flask", "django", "streamlit", "plotly",
]

STATUS_CHOICES = ["All", "Selected", "Waiting List", "Rejected"]

# --------------------------------------------------------------------------
# Resume parsing helpers
# --------------------------------------------------------------------------

def extract_email(text):
    found = re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", text or "")
    return found[0] if found else "applicant@example.com"


def extract_name(text):
    lines = [line.strip() for line in (text or "").splitlines() if line.strip()]
    skip_words = ["email", "phone", "linkedin", "github", "resume", "curriculum", "address"]
    for line in lines[:12]:
        low = line.lower()
        if any(word in low for word in skip_words):
            continue
        clean = re.sub(r"[^A-Za-z .]", "", line).strip()
        if 1 <= len(clean.split()) <= 5:
            return clean.title()
    return "Applicant"


def extract_links(text):
    links = re.findall(
        r"(https?://[^\s)>,]+|www\.[^\s)>,]+|github\.com/[^\s)>,]+|linkedin\.com/[^\s)>,]+)",
        text or "",
        flags=re.I,
    )
    cleaned = []
    for link in links:
        link = link.strip().rstrip(".,;")
        if link not in cleaned:
            cleaned.append(link)
    return cleaned[:6]


def read_resume(filepath):
    """Read resume text from a filepath (Gradio gives us a temp file path)."""
    if not filepath:
        return ""

    filename = filepath.lower()

    with open(filepath, "rb") as f:
        data = f.read()

    if filename.endswith(".txt"):
        return data.decode("utf-8", errors="ignore")

    if filename.endswith(".pdf"):
        import PyPDF2

        reader = PyPDF2.PdfReader(io.BytesIO(data))
        text_parts = []
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
        return "\n".join(text_parts)

    if filename.endswith(".docx"):
        import docx

        document = docx.Document(io.BytesIO(data))
        return "\n".join([p.text for p in document.paragraphs])

    if filename.endswith((".jpg", ".jpeg", ".png")):
        try:
            from PIL import Image
            import pytesseract

            image = Image.open(io.BytesIO(data))
            return pytesseract.image_to_string(image)
        except Exception:
            return (
                f"Image resume uploaded: {os.path.basename(filepath)}. "
                "OCR is not installed. Use PDF, DOCX, or TXT for full text extraction."
            )

    return ""


def split_terms(value):
    parts = re.split(r"[,;\n/|]+", value or "")
    return [part.strip().lower() for part in parts if part.strip()]


def detect_skills(text):
    low = (text or "").lower()
    found = [skill.title() for skill in SKILL_BANK if skill in low]
    return sorted(set(found))


def extract_projects(text):
    lines = [line.strip() for line in (text or "").splitlines() if line.strip()]
    projects = []
    capture = False
    for line in lines:
        low = line.lower()
        if any(word in low for word in ["project", "projects", "portfolio", "github"]):
            capture = True
            clean = re.sub(r"\s+", " ", line).strip("•-: ")
            if len(clean) > 5 and clean not in projects:
                projects.append(clean)
            continue
        if capture:
            if any(stop in low for stop in ["education", "certification", "skills", "experience", "declaration"]):
                capture = False
                continue
            clean = re.sub(r"\s+", " ", line).strip("•-: ")
            if 8 <= len(clean) <= 150 and clean not in projects:
                projects.append(clean)
        if len(projects) >= 5:
            break
    return projects[:5]


def ai_prediction(text, role, skills, requirements, qualifications, experience):
    resume_text = (text or "").lower()
    required_skills = split_terms(skills)
    requirement_terms = split_terms(requirements)
    qualification_terms = split_terms(qualifications)

    matched, missing = [], []
    for skill in required_skills:
        if skill in resume_text:
            matched.append(skill.title())
        else:
            missing.append(skill.title())

    applicant_skills = detect_skills(text)
    if not applicant_skills and matched:
        applicant_skills = matched

    positive_words = [
        "project", "internship", "experience", "certificate", "certification",
        "developed", "created", "built", "managed", "analysis", "dashboard",
        "report", "team", "communication", "presentation", "portfolio",
    ]

    signal_hits = sum(1 for word in positive_words if word in resume_text)
    skill_ratio = len(matched) / max(len(required_skills), 1)
    requirement_hits = sum(1 for word in requirement_terms if word in resume_text)
    qualification_hits = sum(1 for word in qualification_terms if word in resume_text)

    score = 45
    score += int(skill_ratio * 35)
    score += min(8, requirement_hits * 2)
    score += min(7, signal_hits)
    score += min(5, qualification_hits * 2)

    if "fresher" in experience.lower() or "<0" in experience:
        score += 3
    elif any(word in resume_text for word in ["experience", "internship", "worked", "year"]):
        score += 4

    if len(required_skills) > 0 and len(matched) == 0:
        score = min(score, 58)
    if len(required_skills) >= 3 and skill_ratio < 0.35:
        score = min(score, 66)

    score = max(45, min(score, 94))

    if score >= 78:
        decision, fit_level = "Selected", "Strong Fit"
        reason = "AI predicts this applicant is suitable because the resume shows relevant skill alignment, role fit, projects, or positive profile signals."
    elif score >= 62:
        decision, fit_level = "Waiting List", "Potential Fit"
        reason = "AI predicts this applicant has partial alignment. HR can keep this candidate in waiting list or review manually."
    else:
        decision, fit_level = "Rejected", "Needs Review"
        reason = "AI predicts this applicant is currently not the strongest match for the selected role requirements."

    return {
        "score": score,
        "decision": decision,
        "fit_level": fit_level,
        "applicant_skills": applicant_skills,
        "missing_skills": missing[:8],
        "reason": reason,
    }


def email_content(name, role, decision):
    if decision == "Selected":
        return f"""Dear {name},

Thank you for your interest in the {role} position.

We are pleased to inform you that your profile has been shortlisted for the next stage of our recruitment process. Our team found your background relevant to the requirements of this role.

Our HR team will contact you soon with further details regarding the next steps.

Best regards,
HR Recruitment Team"""

    if decision == "Waiting List":
        return f"""Dear {name},

Thank you for applying for the {role} position.

After reviewing your application, we found your profile suitable for future consideration. At this stage, we have placed your application on our waiting list.

If a suitable opportunity becomes available, our HR team will contact you regarding the next steps.

Best regards,
HR Recruitment Team"""

    return f"""Dear {name},

Thank you for applying for the {role} position.

We appreciate the time and effort you invested in your application. After careful review, we regret to inform you that we will not be moving forward with your application at this stage.

We encourage you to apply again for future opportunities that match your profile.

Best regards,
HR Recruitment Team"""


# --------------------------------------------------------------------------
# Excel export helpers
# --------------------------------------------------------------------------

def excel_style_header(sheet):
    from openpyxl.styles import Font, PatternFill, Alignment

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.alignment = Alignment(horizontal="center")


def auto_width(sheet):
    for col in sheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            max_len = max(max_len, len(str(cell.value or "")))
        sheet.column_dimensions[col_letter].width = max_len + 4


def safe_save_workbook(workbook, path):
    try:
        workbook.save(path)
        return path
    except PermissionError:
        base, ext = os.path.splitext(path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback_path = f"{base}_{timestamp}{ext}"
        workbook.save(fallback_path)
        return fallback_path


def _write_rows_excel(rows, path, sheet_name):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    headers = ["Name", "Email", "Role", "AI Confidence", "Status", "Experience", "Qualification", "Date"]
    sheet.append(headers)
    excel_style_header(sheet)

    for row in rows:
        sheet.append([
            row["name"], row["email"], row["role"], row["score"], row["status"],
            row["experience"], row["qualification"], row["date"],
        ])

    auto_width(sheet)
    return safe_save_workbook(workbook, path)


def write_candidates_excel(status, path, sheet_name):
    conn = db()
    rows = conn.execute(
        "SELECT name,email,role,score,status,experience,qualification,date FROM candidates WHERE status=? ORDER BY id DESC",
        (status,),
    ).fetchall()
    conn.close()
    return _write_rows_excel(rows, path, sheet_name)


def write_all_candidates_excel():
    conn = db()
    rows = conn.execute(
        "SELECT name,email,role,score,status,experience,qualification,date FROM candidates ORDER BY id DESC"
    ).fetchall()
    conn.close()
    return _write_rows_excel(rows, ALL_EXCEL_PATH, "All Candidates")


def write_all_excel_files():
    return {
        "selected": write_candidates_excel("Selected", SELECTED_EXCEL_PATH, "Selected Candidates"),
        "waiting": write_candidates_excel("Waiting List", WAITING_EXCEL_PATH, "Waiting List Candidates"),
        "rejected": write_candidates_excel("Rejected", REJECTED_EXCEL_PATH, "Rejected Candidates"),
        "all": write_all_candidates_excel(),
    }


# --------------------------------------------------------------------------
# Candidate table helpers (for Gradio dataframe display)
# --------------------------------------------------------------------------
CANDIDATE_COLS = ["ID", "Name", "Email", "Role", "Score", "Status", "Fit Level",
                  "Experience", "Qualification", "Applicant Skills", "Missing Skills", "Date"]


def fetch_candidates(status_filter="All", role_filter=""):
    conn = db()
    rows = conn.execute("SELECT * FROM candidates ORDER BY id DESC").fetchall()
    conn.close()
    data = []
    for row in rows:
        if status_filter != "All" and row["status"] != status_filter:
            continue
        if role_filter and row["role"] != role_filter:
            continue
        data.append([
            row["id"], row["name"], row["email"], row["role"], row["score"], row["status"],
            row["fit_level"], row["experience"], row["qualification"],
            row["applicant_skills"], row["missing_skills"], row["date"],
        ])
    return pd.DataFrame(data, columns=CANDIDATE_COLS)


def refresh_candidates(status_filter, role_filter):
    return fetch_candidates(status_filter, role_filter or "")


def dashboard_stats():
    conn = db()
    rows = conn.execute("SELECT status, score FROM candidates").fetchall()
    conn.close()
    total = len(rows)
    selected = sum(1 for r in rows if r["status"] == "Selected")
    waiting = sum(1 for r in rows if r["status"] == "Waiting List")
    rejected = sum(1 for r in rows if r["status"] in ("Rejected", "Not Selected"))
    avg_score = round(sum(r["score"] or 0 for r in rows) / total, 1) if total else 0

    md = f"""
### 📊 Pipeline Overview
| Metric | Value |
|---|---|
| Total Applicants | **{total}** |
| ✅ Selected | **{selected}** |
| ⏳ Waiting List | **{waiting}** |
| ❌ Rejected | **{rejected}** |
| Average AI Confidence Score | **{avg_score}** |
"""
    return md


# --------------------------------------------------------------------------
# Core actions
# --------------------------------------------------------------------------

def analyze_resume(resume_file, role, department, skills, requirements,
                    qualification_type, qualifications, experience,
                    minimum_projects, nlp_brief):
    if not resume_file:
        return (
            "⚠️ Please upload a resume first.", "", "", "", "",
            dashboard_stats(), fetch_candidates(),
        )

    text = read_resume(resume_file)
    name = extract_name(text)
    email = extract_email(text)
    projects = extract_projects(text)
    links = extract_links(text)

    combined_qualifications = f"{qualification_type}, {qualifications}".strip(", ")
    combined_requirements = requirements
    if nlp_brief:
        combined_requirements = f"{requirements}, {nlp_brief}"

    ai = ai_prediction(text, role, skills, combined_requirements, combined_qualifications, experience)
    decision = ai["decision"]
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    body = email_content(name, role, decision)

    conn = db()
    conn.execute(
        """
        INSERT INTO candidates(
            name,email,role,score,decision,status,date,email_body,
            experience,qualification,applicant_skills,missing_skills,reason,fit_level,projects,project_links
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            name, email, role, ai["score"], decision, decision, now, body,
            experience, combined_qualifications, ", ".join(ai["applicant_skills"]),
            ", ".join(ai["missing_skills"]), ai["reason"], ai["fit_level"],
            " || ".join(projects), " || ".join(links),
        ),
    )
    conn.commit()
    conn.close()

    write_all_excel_files()

    summary = f"""
### 🎯 Screening Result — {name}

| Field | Value |
|---|---|
| Email | {email} |
| Role | {role} |
| **AI Confidence Score** | **{ai['score']} / 100** |
| **Decision** | **{decision}** ({ai['fit_level']}) |
| Experience | {experience} |
| Qualification | {combined_qualifications} |

**Matched / Detected Skills:** {", ".join(ai["applicant_skills"]) or "None detected"}

**Missing Skills:** {", ".join(ai["missing_skills"]) or "None"}

**AI Reasoning:** {ai["reason"]}

**Projects found:** {"; ".join(projects) or "None detected"}

**Links found:** {", ".join(links) or "None detected"}
"""

    return (
        summary, name, email, decision, body,
        dashboard_stats(), fetch_candidates(),
    )


def change_status(candidate_id, new_status):
    if not candidate_id:
        return "⚠️ Enter a candidate ID.", dashboard_stats(), fetch_candidates()
    try:
        candidate_id = int(candidate_id)
    except ValueError:
        return "⚠️ Candidate ID must be a number.", dashboard_stats(), fetch_candidates()

    conn = db()
    conn.execute("UPDATE candidates SET status=?, decision=? WHERE id=?", (new_status, new_status, candidate_id))
    conn.commit()
    conn.close()
    write_all_excel_files()
    return f"✅ Candidate #{candidate_id} moved to **{new_status}**.", dashboard_stats(), fetch_candidates()


def delete_candidate(candidate_id):
    if not candidate_id:
        return "⚠️ Enter a candidate ID.", dashboard_stats(), fetch_candidates()
    try:
        candidate_id = int(candidate_id)
    except ValueError:
        return "⚠️ Candidate ID must be a number.", dashboard_stats(), fetch_candidates()

    conn = db()
    conn.execute("DELETE FROM candidates WHERE id=?", (candidate_id,))
    conn.commit()
    conn.close()
    write_all_excel_files()
    return f"🗑️ Candidate #{candidate_id} deleted.", dashboard_stats(), fetch_candidates()


def export_and_get_file(kind):
    write_all_excel_files()
    paths = {
        "selected": SELECTED_EXCEL_PATH,
        "waiting": WAITING_EXCEL_PATH,
        "rejected": REJECTED_EXCEL_PATH,
        "all": ALL_EXCEL_PATH,
    }
    return paths[kind]


# ---- Email sending ----

def send_email(sender_email, sender_password, to_email, subject, body):
    if not sender_email or not sender_password or not to_email:
        return "⚠️ Sender email, sender app-password, and recipient email are all required."

    msg = MIMEText(body or "")
    msg["Subject"] = subject or "Regarding your application"
    msg["From"] = sender_email
    msg["To"] = to_email

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, [to_email], msg.as_string())
        server.quit()
        return "✅ Email sent successfully."
    except Exception as error:
        return f"❌ Email failed: {error}"


# ---- AI text generators (rule-based, mirrors original /ai/generate & /ai/interview-questions) ----

def generate_jd(role, department, skills, requirements, qualification_type, qualifications, experience, minimum_projects, nlp_brief):
    skills_list = [s.strip() for s in re.split(r"[,;/|]+", skills) if s.strip()] or ["Communication", "Problem Solving", "Teamwork"]
    requirements_list = [r.strip() for r in re.split(r"[,;/|]+", requirements) if r.strip()] or ["Good communication", "Willingness to learn", "Professional attitude"]
    qualification = f"{qualification_type}, {qualifications}".strip(", ")

    role_lower = (role or "").lower()
    if any(k in role_lower for k in ["marketing", "seo", "social media"]):
        responsibilities = [
            "Assist in planning and executing digital marketing activities.",
            "Support campaign research, content planning, and competitor analysis.",
            "Track basic campaign performance using Excel, analytics tools, or dashboards.",
            "Coordinate with design, content, and business teams for marketing tasks.",
            "Prepare simple reports and insights for weekly review.",
        ]
    elif any(k in role_lower for k in ["data", "analyst", "power bi"]):
        responsibilities = [
            "Collect, clean, and organize data for reporting.",
            "Create dashboards and visual summaries using analytics tools.",
            "Identify trends, gaps, and useful business insights from data.",
            "Support teams with Excel, SQL, Power BI, or reporting tasks.",
            "Document findings clearly for decision-making.",
        ]
    elif any(k in role_lower for k in ["developer", "frontend", "backend", "software"]):
        responsibilities = [
            "Develop and maintain application features based on requirements.",
            "Write clean, readable, and reusable code.",
            "Collaborate with team members to test and debug issues.",
            "Work with APIs, databases, or frontend components as required.",
            "Maintain basic documentation for completed tasks.",
        ]
    elif any(k in role_lower for k in ["hr", "recruit", "talent"]):
        responsibilities = [
            "Assist in screening resumes and maintaining candidate records.",
            "Coordinate interview communication and recruitment updates.",
            "Support HR documentation, reports, and applicant tracking.",
            "Follow up with shortlisted, waiting list, and rejected candidates.",
            "Maintain professional communication with applicants.",
        ]
    else:
        responsibilities = [
            f"Support daily tasks related to the {role} role.",
            "Coordinate with team members to complete assigned work.",
            "Maintain reports, documentation, and task updates.",
            "Communicate clearly with internal stakeholders.",
            "Learn required tools and contribute to ongoing projects.",
        ]

    skill_lines = "\n".join(f"- {s}" for s in skills_list)
    requirement_lines = "\n".join(f"- {r}" for r in requirements_list)
    responsibility_lines = "\n".join(f"- {r}" for r in responsibilities)

    jd = f"""# {role}

## Job Title
{role}

## Department
{department}

## Job Summary
We are looking for a motivated {role} to join the {department} team. The ideal candidate should have a learning mindset, relevant practical skills, and the ability to complete assigned work with professionalism and consistency.

## Key Responsibilities
{responsibility_lines}

## Required Skills
{skill_lines}

## Eligibility Criteria
- Qualification: {qualification}
- Experience: {experience}
- Minimum Projects: {minimum_projects}
- Candidate should be able to demonstrate relevant skills through projects, internships, certifications, or practical work.
- Freshers can apply if they show strong learning ability and role alignment.

## Preferred Qualities
- Clear communication
- Professional attitude
- Ownership mindset
- Ability to work with deadlines
- Curiosity and willingness to learn
- Team collaboration

## Ideal Candidate Profile
The ideal candidate is someone who understands the basics of the role, can learn quickly, communicates clearly, and shows proof of interest through projects, certifications, internships, or portfolio work.

## Selection Criteria
- Skills match: 40%
- Communication and professionalism: 20%
- Projects or practical proof: 20%
- Learning ability and attitude: 20%

## Short LinkedIn Job Post Version
We are hiring for {role} in the {department} department.

If you have skills in {", ".join(skills_list[:4])}, good communication, and a strong willingness to learn, this opportunity is for you.

Apply with your updated resume and relevant project or certification proof.
{"Additional HR brief considered: " + nlp_brief if nlp_brief else ""}
"""
    return jd


def generate_interview_questions(role, skills, experience):
    role = (role or "Selected Role").strip()
    experience = (experience or "Fresher").strip()
    skill_list = [s.strip() for s in re.split(r"[,;/|]+", skills or "") if s.strip()] or ["Communication", "Problem Solving"]
    skill_text = ", ".join(skill_list[:6])
    skill_bullets = "\n".join(f"- {s}" for s in skill_list)

    return f"""# Interview Questions for {role}

## Candidate Level
Experience Required: {experience}

## Core Skill Areas
{skill_bullets}

## Technical / Role-Based Questions
1. Explain your understanding of the {role} role.
2. Which of these skills have you used before: {skill_text}?
3. Describe one project, internship, or task that is related to this role.
4. What tools or platforms would you use to complete work in this role?
5. How would you approach a new assignment if you had limited guidance?

## Skill-Based Questions
1. Pick one skill from this list and explain how you have used it: {skill_text}.
2. What is one challenge you faced while learning or applying these skills?
3. How do you keep improving your practical skills?
4. Can you show or explain proof of work related to this role?

## HR Questions
1. Tell me about yourself.
2. Why are you interested in this role?
3. What are your strengths that match this position?
4. What is one weakness you are currently improving?
5. Why should we shortlist you for this opportunity?

## Situational Questions
1. What would you do if you were assigned a task you do not know how to complete?
2. How would you handle feedback from a senior or team lead?
3. How would you manage multiple tasks with the same deadline?
4. What would you do if your work had an error close to submission time?
5. How would you communicate delays or blockers to your manager?

## Practical Task
Give the candidate a small task related to {role} using these skills: {skill_text}.

Example Task:
Prepare a short work sample, mini report, project outline, dashboard, campaign idea, code module, or analysis based on the role requirements. Ask the candidate to explain their approach, tools used, and final output.

## Evaluation Criteria
- Role understanding: 20%
- Skill confidence: 30%
- Practical proof/project explanation: 25%
- Communication: 15%
- Learning attitude: 10%
"""


# --------------------------------------------------------------------------
# UI
# --------------------------------------------------------------------------
# --------------------------------------------------------------------------
# UI
# --------------------------------------------------------------------------

CUSTOM_CSS = """
:root{
  --bg:#f4f8fb;
  --panel:#ffffff;
  --ink:#0f172a;
  --muted:#64748b;
  --line:#dbe4ee;
  --teal:#0f766e;
  --teal2:#14b8a6;
  --nav:#08111f;
  --nav2:#101c2f;
}
html,body{background:var(--bg)!important}
.gradio-container{
  max-width:1600px!important;
  margin:0 auto!important;
  padding:0!important;
  background:var(--bg)!important;
  font-family:Inter,system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif!important;
  color:var(--ink)!important;
}
footer{display:none!important}
button{font-weight:800!important}
input,textarea,select{border-radius:12px!important}
.block,.form,.gr-box{border-radius:16px!important}

#login-page{
  min-height:100vh;
  padding:28px;
  background:
    radial-gradient(circle at 10% 10%,rgba(20,184,166,.24),transparent 28%),
    radial-gradient(circle at 88% 18%,rgba(56,189,248,.18),transparent 28%),
    linear-gradient(135deg,#ecfeff,#f8fafc 52%,#eef2ff);
}
.login-wrap{
  min-height:calc(100vh - 56px);
  display:grid;
  grid-template-columns:1.12fr .88fr;
  gap:24px;
  align-items:stretch;
}
.login-hero,.login-card{
  background:rgba(255,255,255,.92);
  border:1px solid rgba(255,255,255,.8);
  border-radius:30px;
  box-shadow:0 24px 70px rgba(15,23,42,.11);
}
.login-hero{
  padding:48px;
  display:flex;
  flex-direction:column;
  justify-content:center;
}
.brand{
  display:flex;align-items:center;gap:14px;margin-bottom:36px
}
.brand-logo{
  width:52px;height:52px;border-radius:17px;display:grid;place-items:center;
  background:linear-gradient(135deg,#0f766e,#14b8a6);color:#fff;font-weight:950;font-size:20px;
  box-shadow:0 14px 30px rgba(13,148,136,.25)
}
.brand-name{font-size:25px;font-weight:950;letter-spacing:-.03em}
.brand-tag{font-size:11px;text-transform:uppercase;letter-spacing:.18em;color:#0f766e;font-weight:900;margin-top:4px}
.login-hero h1{font-size:clamp(44px,6vw,78px);line-height:.96;letter-spacing:-.06em;margin:0 0 18px;font-weight:950}
.login-hero h1 span{color:#0f766e}
.login-hero p{font-size:18px;line-height:1.75;color:#526176;max-width:720px;font-weight:650}
.hero-features{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-top:30px}
.hero-feature{padding:16px;border-radius:18px;background:#f8fafc;border:1px solid var(--line)}
.hero-feature b{display:block;font-size:20px}
.hero-feature span{font-size:12px;color:var(--muted);font-weight:800}
.login-card{padding:30px;display:flex;flex-direction:column;justify-content:center}
.ai-preview{
  background:linear-gradient(145deg,#07111f,#10233b);
  color:#fff;border-radius:24px;padding:22px;margin-bottom:22px;
}
.preview-top{display:flex;justify-content:space-between;align-items:center}
.preview-score{width:72px;height:72px;border-radius:22px;display:grid;place-items:center;background:#34d399;color:#07111f;font-size:23px;font-weight:950}
.preview-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-top:16px}
.preview-cell{background:rgba(255,255,255,.08);border-radius:14px;text-align:center;padding:10px}
.preview-cell small{color:#cbd5e1}.preview-cell b{display:block;margin-top:3px}

#app-page{min-height:100vh}
.app-shell{display:grid;grid-template-columns:270px 1fr;min-height:100vh}
.sidebar{
  background:linear-gradient(180deg,var(--nav),var(--nav2));
  color:#fff;padding:22px 18px;position:sticky;top:0;height:100vh;
}
.side-brand{display:flex;align-items:center;gap:12px;padding:5px 8px 24px}
.side-brand .logo{width:44px;height:44px;border-radius:15px;display:grid;place-items:center;background:linear-gradient(135deg,#14b8a6,#38bdf8);font-weight:950}
.side-brand b{font-size:21px}.side-brand small{display:block;color:#94a3b8;margin-top:3px;font-weight:700}
.side-nav-label{color:#64748b;font-size:11px;font-weight:900;letter-spacing:.16em;text-transform:uppercase;margin:20px 10px 10px}
.nav-button button{
  width:100%!important;text-align:left!important;justify-content:flex-start!important;
  border-radius:14px!important;border:0!important;background:transparent!important;color:#cbd5e1!important;
  padding:12px 14px!important;margin:4px 0!important
}
.nav-button button:hover{background:rgba(255,255,255,.08)!important;color:#fff!important}
.nav-button.active button{background:linear-gradient(90deg,#0f766e,#0d9488)!important;color:#fff!important}
.sidebar-bottom{position:absolute;left:18px;right:18px;bottom:20px}
.logout-btn button{width:100%!important;background:rgba(255,255,255,.08)!important;color:#fff!important;border:1px solid rgba(255,255,255,.08)!important;border-radius:14px!important}

.main-area{padding:24px 28px 40px;min-width:0}
.topbar{
  display:flex;justify-content:space-between;align-items:center;gap:20px;
  background:#fff;border:1px solid var(--line);border-radius:22px;padding:20px 22px;margin-bottom:20px;
  box-shadow:0 12px 32px rgba(15,23,42,.06)
}
.topbar h1{margin:0;font-size:31px;letter-spacing:-.04em;font-weight:950}
.topbar p{margin:6px 0 0;color:var(--muted);font-weight:700}
.user-chip{display:flex;align-items:center;gap:10px;background:#f8fafc;border:1px solid var(--line);padding:9px 13px;border-radius:999px;font-weight:850}
.user-dot{width:34px;height:34px;border-radius:50%;display:grid;place-items:center;background:#ccfbf1;color:#0f766e;font-weight:950}

.kpis{display:grid;grid-template-columns:repeat(5,1fr);gap:14px;margin-bottom:20px}
.kpi{background:#fff;border:1px solid var(--line);border-radius:20px;padding:18px;box-shadow:0 10px 28px rgba(15,23,42,.05)}
.kpi-icon{width:42px;height:42px;border-radius:14px;display:grid;place-items:center;background:#ccfbf1;color:#0f766e;font-size:19px}
.kpi strong{display:block;font-size:27px;margin-top:13px;font-weight:950}
.kpi span{display:block;color:var(--muted);font-size:12px;font-weight:800;margin-top:3px}

.content-card{background:#fff;border:1px solid var(--line);border-radius:22px;padding:22px;box-shadow:0 12px 32px rgba(15,23,42,.05);margin-bottom:18px}
.section-head{display:flex;justify-content:space-between;gap:16px;align-items:flex-start;margin-bottom:18px}
.section-head h2{margin:0;font-size:24px;letter-spacing:-.03em;font-weight:950}
.section-head p{margin:6px 0 0;color:var(--muted);font-weight:650}
.primary-btn button{background:linear-gradient(90deg,#0f766e,#14b8a6)!important;color:#fff!important;border:0!important;border-radius:14px!important;box-shadow:0 10px 24px rgba(13,148,136,.20)!important}
.danger-btn button{background:#dc2626!important;color:#fff!important;border:0!important;border-radius:14px!important}
.secondary-btn button{border-radius:14px!important}
.dark-result{
  background:linear-gradient(160deg,#07111f,#10233b)!important;
  color:#fff!important;border-radius:20px!important;padding:20px!important;min-height:330px
}
.dark-result h1,.dark-result h2,.dark-result h3,.dark-result strong,.dark-result table{color:#fff!important}
.page{animation:fade .25s ease both}
@keyframes fade{from{opacity:.4;transform:translateY(6px)}to{opacity:1;transform:none}}
@media(max-width:1100px){
  .app-shell{grid-template-columns:220px 1fr}.kpis{grid-template-columns:repeat(3,1fr)}
  .login-wrap{grid-template-columns:1fr}.hero-features{grid-template-columns:repeat(2,1fr)}
}
@media(max-width:760px){
  .app-shell{display:block}.sidebar{height:auto;position:relative}.sidebar-bottom{position:static;margin-top:18px}
  .kpis{grid-template-columns:repeat(2,1fr)}.main-area{padding:14px}.login-hero{padding:28px}
}
"""

def do_login(email, password):
    if not email or not password:
        return gr.update(), gr.update(), "⚠️ Enter your email and password.", "HR Recruiter"
    name = email.split("@")[0].replace(".", " ").replace("_", " ").title()
    return gr.update(visible=False), gr.update(visible=True), "", name

def do_signup(name, email, password, role):
    if not name or not email or not password:
        return gr.update(), gr.update(), "⚠️ Complete all required fields.", name or "HR Recruiter"
    return gr.update(visible=False), gr.update(visible=True), "", name

def do_logout():
    return gr.update(visible=True), gr.update(visible=False)

def page_state(target):
    pages = ["dashboard","screening","pipeline","jd","interview","email"]
    return [gr.update(visible=(p == target)) for p in pages]

with gr.Blocks(title="HR Mate — AI Recruitment Platform", css=CUSTOM_CSS) as demo:

    # LOGIN PAGE
    with gr.Column(visible=True, elem_id="login-page") as login_page:
        with gr.Row(elem_classes=["login-wrap"]):
            gr.HTML("""
            <section class="login-hero">
              <div class="brand">
                <div class="brand-logo">HM</div>
                <div><div class="brand-name">HR Mate</div><div class="brand-tag">AI Recruitment Platform</div></div>
              </div>
              <h1>Hire smarter with <span>AI-powered HR</span></h1>
              <p>Screen resumes, generate job descriptions, build interview kits, manage candidates and automate outcome communication from one professional workspace.</p>
              <div class="hero-features">
                <div class="hero-feature"><b>15 sec</b><span>Resume screening</span></div>
                <div class="hero-feature"><b>AI Score</b><span>Candidate fit analysis</span></div>
                <div class="hero-feature"><b>Excel</b><span>Pipeline exports</span></div>
                <div class="hero-feature"><b>Email</b><span>Outcome automation</span></div>
              </div>
            </section>
            """)
            with gr.Column(elem_classes=["login-card"]):
                gr.HTML("""
                <div class="ai-preview">
                  <div class="preview-top">
                    <div><small style="color:#a5f3fc;font-weight:800">Live AI candidate preview</small><h2 style="margin:5px 0 0">Strong fit identified</h2></div>
                    <div class="preview-score">92%</div>
                  </div>
                  <div class="preview-grid">
                    <div class="preview-cell"><small>Status</small><b style="color:#6ee7b7">Selected</b></div>
                    <div class="preview-cell"><small>Role</small><b>Analyst</b></div>
                    <div class="preview-cell"><small>Skills</small><b>8 / 10</b></div>
                  </div>
                </div>
                <h2 style="font-size:28px;margin:0;font-weight:950">Welcome to HR Mate</h2>
                <p style="color:#64748b;font-weight:700;margin:7px 0 15px">Access your recruitment workspace.</p>
                """)
                with gr.Tabs():
                    with gr.Tab("Login"):
                        login_email = gr.Textbox(label="Company Email", placeholder="hr@company.com")
                        login_password = gr.Textbox(label="Password", type="password")
                        login_btn = gr.Button("Login to Workspace", variant="primary", elem_classes=["primary-btn"])
                    with gr.Tab("Sign Up"):
                        signup_name = gr.Textbox(label="Full Name")
                        signup_email = gr.Textbox(label="Company Email")
                        signup_password = gr.Textbox(label="Create Password", type="password")
                        signup_role = gr.Dropdown(["HR Recruiter","Talent Acquisition","HR Manager","Recruitment Lead"], value="HR Recruiter", label="Role")
                        signup_btn = gr.Button("Create Workspace", variant="primary", elem_classes=["primary-btn"])
                login_msg = gr.Markdown()

    # APP PAGE
    with gr.Column(visible=False, elem_id="app-page") as app_page:
        with gr.Row(elem_classes=["app-shell"]):
            with gr.Column(elem_classes=["sidebar"], min_width=220):
                gr.HTML("""
                <div class="side-brand">
                  <div class="logo">HM</div>
                  <div><b>HR Mate</b><small>AI Recruitment</small></div>
                </div>
                <div class="side-nav-label">Workspace</div>
                """)
                nav_dashboard = gr.Button("▦  Dashboard", elem_classes=["nav-button","active"])
                nav_screening = gr.Button("⌕  Resume Screening", elem_classes=["nav-button"])
                nav_pipeline = gr.Button("👥  Applicants & Pipeline", elem_classes=["nav-button"])
                nav_jd = gr.Button("✦  Job Description Generator", elem_classes=["nav-button"])
                nav_interview = gr.Button("🎤  Interview Questions", elem_classes=["nav-button"])
                nav_email = gr.Button("✉  Email Automation", elem_classes=["nav-button"])
                gr.HTML('<div class="side-nav-label">Account</div>')
                logout_btn = gr.Button("↪  Logout", elem_classes=["logout-btn"])

            with gr.Column(elem_classes=["main-area"]):
                current_user = gr.Textbox(value="HR Recruiter", visible=False)

                gr.HTML("""
                <div class="topbar">
                  <div><h1>HR Recruitment Workspace</h1><p>AI-powered screening, hiring automation and candidate intelligence</p></div>
                  <div class="user-chip"><div class="user-dot">HR</div><span>Recruitment Team</span></div>
                </div>
                <div class="kpis">
                  <div class="kpi"><div class="kpi-icon">👥</div><strong>Live</strong><span>Candidate pipeline</span></div>
                  <div class="kpi"><div class="kpi-icon">✓</div><strong>AI</strong><span>Resume decisions</span></div>
                  <div class="kpi"><div class="kpi-icon">⏱</div><strong>Fast</strong><span>Automated screening</span></div>
                  <div class="kpi"><div class="kpi-icon">📊</div><strong>XLSX</strong><span>Recruitment reports</span></div>
                  <div class="kpi"><div class="kpi-icon">✉</div><strong>SMTP</strong><span>Outcome emails</span></div>
                </div>
                """)

                # DASHBOARD
                with gr.Column(visible=True, elem_classes=["page"], elem_id="dashboard") as page_dashboard:
                    gr.HTML("""
                    <div class="content-card">
                      <div class="section-head">
                        <div><h2>Executive Recruitment Dashboard</h2><p>Manage your entire recruitment workflow from one place.</p></div>
                      </div>
                    </div>
                    """)
                    with gr.Row():
                        with gr.Column(elem_classes=["content-card"]):
                            dashboard_md = gr.Markdown(dashboard_stats())
                        with gr.Column(elem_classes=["content-card"]):
                            gr.Markdown("""
### Quick Actions
- Upload and screen a new resume
- Generate a professional job description
- Create role-based interview questions
- Review and update candidate status
- Export candidate records to Excel
- Send candidate outcome emails
""")

                # SCREENING
                with gr.Column(visible=False, elem_classes=["page"], elem_id="screening") as page_screening:
                    gr.HTML('<div class="content-card"><div class="section-head"><div><h2>AI Resume Screening</h2><p>Evaluate candidate resumes against role requirements and generate a fit decision.</p></div></div></div>')
                    with gr.Row():
                        with gr.Column(elem_classes=["content-card"]):
                            role = gr.Dropdown(JOB_ROLES, value=JOB_ROLES[0], label="Role", allow_custom_value=True)
                            department = gr.Textbox(value="Marketing", label="Department")
                            skills = gr.Textbox(value="Python, Power BI, Excel, Communication", label="Required Skills")
                            requirements = gr.Textbox(value="Communication skills, enthusiasm, project experience", label="Key Requirements")
                            qualification_type = gr.Dropdown(QUALIFICATION_OPTIONS, value="Any Degree", label="Qualification Type")
                            qualifications = gr.Textbox(value="MBA, B.Tech, relevant certifications", label="Qualification Details")
                            experience = gr.Dropdown(EXPERIENCE_OPTIONS, value=EXPERIENCE_OPTIONS[0], label="Experience Required")
                            minimum_projects = gr.Textbox(value="0", label="Minimum Projects")
                            nlp_brief = gr.Textbox(label="Additional HR Brief", lines=3)
                            resume_file = gr.File(label="Upload Resume", file_types=[".pdf",".docx",".txt",".png",".jpg",".jpeg"])
                            analyze_btn = gr.Button("Analyze Candidate Resume", variant="primary", elem_classes=["primary-btn"])
                        with gr.Column(elem_classes=["content-card"]):
                            result_md = gr.Markdown("### Candidate intelligence will appear here\nUpload a resume and click **Analyze Candidate Resume**.", elem_classes=["dark-result"])
                            with gr.Accordion("Candidate Details & Outcome Email", open=True):
                                cand_name = gr.Textbox(label="Candidate Name", interactive=False)
                                cand_email = gr.Textbox(label="Candidate Email", interactive=False)
                                cand_decision = gr.Textbox(label="AI Decision", interactive=False)
                                email_body_box = gr.Textbox(label="Outcome Email Draft", lines=10)

                # PIPELINE
                with gr.Column(visible=False, elem_classes=["page"], elem_id="pipeline") as page_pipeline:
                    gr.HTML('<div class="content-card"><div class="section-head"><div><h2>Applicants & Pipeline</h2><p>Track, filter, update and export every candidate record.</p></div></div></div>')
                    dash_md = gr.Markdown(dashboard_stats(), elem_classes=["content-card"])
                    with gr.Row(elem_classes=["content-card"]):
                        status_filter = gr.Dropdown(STATUS_CHOICES, value="All", label="Filter by Status")
                        role_filter = gr.Dropdown(["All"] + JOB_ROLES, value="All", label="Filter by Role")
                        refresh_btn = gr.Button("Refresh Pipeline", elem_classes=["secondary-btn"])
                    with gr.Column(elem_classes=["content-card"]):
                        candidates_df = gr.Dataframe(value=fetch_candidates(), headers=CANDIDATE_COLS, wrap=True, interactive=False)
                    with gr.Column(elem_classes=["content-card"]):
                        gr.Markdown("### Manage Candidate")
                        candidate_id_box = gr.Textbox(label="Candidate ID")
                        action_msg = gr.Markdown()
                        with gr.Row():
                            select_btn = gr.Button("Mark Selected", elem_classes=["primary-btn"])
                            waitlist_btn = gr.Button("Move to Waiting List")
                            reject_btn = gr.Button("Mark Rejected")
                            delete_btn = gr.Button("Delete Candidate", elem_classes=["danger-btn"])
                    with gr.Column(elem_classes=["content-card"]):
                        gr.Markdown("### Excel Exports")
                        with gr.Row():
                            export_selected_btn = gr.Button("Download Selected")
                            export_waiting_btn = gr.Button("Download Waiting")
                            export_rejected_btn = gr.Button("Download Rejected")
                            export_all_btn = gr.Button("Download All")
                        export_file = gr.File(label="Exported File", interactive=False)

                # JD
                with gr.Column(visible=False, elem_classes=["page"], elem_id="jd") as page_jd:
                    gr.HTML('<div class="content-card"><div class="section-head"><div><h2>Job Description Generator</h2><p>Create a complete and professional JD using role requirements.</p></div></div></div>')
                    with gr.Row():
                        with gr.Column(elem_classes=["content-card"]):
                            jd_role = gr.Dropdown(JOB_ROLES, value=JOB_ROLES[0], label="Role", allow_custom_value=True)
                            jd_department = gr.Textbox(value="Marketing", label="Department")
                            jd_skills = gr.Textbox(value="Python, Power BI, Excel, Communication", label="Required Skills")
                            jd_requirements = gr.Textbox(value="Communication skills, enthusiasm, project experience", label="Key Requirements")
                            jd_qual_type = gr.Dropdown(QUALIFICATION_OPTIONS, value="Any Degree", label="Qualification Type")
                            jd_qualifications = gr.Textbox(value="MBA, B.Tech, relevant certifications", label="Qualification Details")
                            jd_experience = gr.Dropdown(EXPERIENCE_OPTIONS, value=EXPERIENCE_OPTIONS[0], label="Experience Required")
                            jd_min_projects = gr.Textbox(value="0", label="Minimum Projects")
                            jd_brief = gr.Textbox(label="Additional HR Brief", lines=3)
                            jd_btn = gr.Button("Generate Professional JD", elem_classes=["primary-btn"])
                        with gr.Column(elem_classes=["content-card"]):
                            jd_output = gr.Markdown("### Generated job description will appear here", elem_classes=["dark-result"])

                # INTERVIEW
                with gr.Column(visible=False, elem_classes=["page"], elem_id="interview") as page_interview:
                    gr.HTML('<div class="content-card"><div class="section-head"><div><h2>Interview Question Generator</h2><p>Generate technical, HR, situational and practical questions.</p></div></div></div>')
                    with gr.Row():
                        with gr.Column(elem_classes=["content-card"]):
                            iq_role = gr.Dropdown(JOB_ROLES, value=JOB_ROLES[0], label="Role", allow_custom_value=True)
                            iq_skills = gr.Textbox(value="Python, Power BI, Excel, Communication", label="Core Skills")
                            iq_experience = gr.Dropdown(EXPERIENCE_OPTIONS, value=EXPERIENCE_OPTIONS[0], label="Candidate Level")
                            iq_btn = gr.Button("Generate Interview Kit", elem_classes=["primary-btn"])
                        with gr.Column(elem_classes=["content-card"]):
                            iq_output = gr.Markdown("### Interview kit will appear here", elem_classes=["dark-result"])

                # EMAIL
                with gr.Column(visible=False, elem_classes=["page"], elem_id="email") as page_email:
                    gr.HTML('<div class="content-card"><div class="section-head"><div><h2>Email Automation</h2><p>Send candidate outcome emails using Gmail SMTP.</p></div></div></div>')
                    with gr.Row():
                        with gr.Column(elem_classes=["content-card"]):
                            sender_email = gr.Textbox(label="HR Gmail Address")
                            sender_password = gr.Textbox(label="Gmail App Password", type="password")
                            to_email = gr.Textbox(label="Candidate Email")
                            subject = gr.Textbox(label="Subject", value="Update on your application")
                            body = gr.Textbox(label="Email Body", lines=12)
                            send_btn = gr.Button("Send Candidate Email", elem_classes=["primary-btn"])
                        with gr.Column(elem_classes=["content-card"]):
                            gr.Markdown("""
### Gmail App Password Setup
Use a Gmail App Password instead of your normal Gmail password.

1. Enable 2-Step Verification.
2. Generate an App Password.
3. Paste the 16-character password.
4. Review the recipient and draft.
5. Send the email.
""")
                            send_result = gr.Markdown()

    # Login / logout
    login_btn.click(do_login, [login_email, login_password], [login_page, app_page, login_msg, current_user])
    signup_btn.click(do_signup, [signup_name, signup_email, signup_password, signup_role], [login_page, app_page, login_msg, current_user])
    logout_btn.click(do_logout, outputs=[login_page, app_page])

    # Navigation
    all_pages = [page_dashboard,page_screening,page_pipeline,page_jd,page_interview,page_email]
    nav_dashboard.click(lambda: page_state("dashboard"), outputs=all_pages)
    nav_screening.click(lambda: page_state("screening"), outputs=all_pages)
    nav_pipeline.click(lambda: page_state("pipeline"), outputs=all_pages)
    nav_jd.click(lambda: page_state("jd"), outputs=all_pages)
    nav_interview.click(lambda: page_state("interview"), outputs=all_pages)
    nav_email.click(lambda: page_state("email"), outputs=all_pages)

    # Core wiring
    analyze_btn.click(
        analyze_resume,
        inputs=[resume_file, role, department, skills, requirements, qualification_type,
                qualifications, experience, minimum_projects, nlp_brief],
        outputs=[result_md, cand_name, cand_email, cand_decision, email_body_box, dash_md, candidates_df],
    )
    refresh_btn.click(refresh_candidates, [status_filter, role_filter], [candidates_df])
    status_filter.change(refresh_candidates, [status_filter, role_filter], [candidates_df])
    role_filter.change(refresh_candidates, [status_filter, role_filter], [candidates_df])

    select_btn.click(lambda cid: change_status(cid, "Selected"), [candidate_id_box], [action_msg, dash_md, candidates_df])
    waitlist_btn.click(lambda cid: change_status(cid, "Waiting List"), [candidate_id_box], [action_msg, dash_md, candidates_df])
    reject_btn.click(lambda cid: change_status(cid, "Rejected"), [candidate_id_box], [action_msg, dash_md, candidates_df])
    delete_btn.click(delete_candidate, [candidate_id_box], [action_msg, dash_md, candidates_df])

    export_selected_btn.click(lambda: export_and_get_file("selected"), outputs=[export_file])
    export_waiting_btn.click(lambda: export_and_get_file("waiting"), outputs=[export_file])
    export_rejected_btn.click(lambda: export_and_get_file("rejected"), outputs=[export_file])
    export_all_btn.click(lambda: export_and_get_file("all"), outputs=[export_file])

    jd_btn.click(
        generate_jd,
        [jd_role, jd_department, jd_skills, jd_requirements, jd_qual_type, jd_qualifications, jd_experience, jd_min_projects, jd_brief],
        [jd_output],
    )
    iq_btn.click(generate_interview_questions, [iq_role, iq_skills, iq_experience], [iq_output])
    send_btn.click(send_email, [sender_email, sender_password, to_email, subject, body], [send_result])

    cand_email.change(lambda v: v, [cand_email], [to_email])
    email_body_box.change(lambda v: v, [email_body_box], [body])
    cand_name.change(lambda n, d: f"Update on your application ({d})" if d else "Update on your application",
                     [cand_name, cand_decision], [subject])

if __name__ == "__main__":
    demo.queue().launch(
        server_name="0.0.0.0",
        server_port=int(os.environ.get("PORT", 7860)),
    )
